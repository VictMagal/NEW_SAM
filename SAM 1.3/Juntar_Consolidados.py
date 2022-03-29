import sys
from PyQt5 import uic, QtWidgets
from PyQt5.QtWidgets import QCompleter
import openpyxl
import pandas as pd
import numpy as np

class load_arquivos:
        def __init__(self):
            self.arquivo_fornecedor = []
            self.list_fornecedor = []
            self.wb_new = openpyxl.Workbook()
            self.ws_new = self.wb_new.active
            
            
        def ler_arquivo_fornecedor(self):
            
            fornecedor = QtWidgets.QFileDialog.getOpenFileNames()[0]
            self.list_fornecedor = fornecedor
            
            print('Load file fornecedor complete!')
            print('list_fornecedor:', self.list_fornecedor)

        def juntar_arquivos(self):
            k = 0
            ultima_linha_ws_new = 1
            
            while k<= len(self.list_fornecedor)-1:
                self.wb = openpyxl.load_workbook(self.list_fornecedor[k])
                self.ws = self.wb ['Worksheet']
                
                is_data = True
                count_row_ws = 1
                while is_data:
                    count_row_ws += 1
                    data =  self.ws.cell (row = count_row_ws, column = 1).value
                    if data == None:
                        is_data = False
                count_row_ws -=1
                
                is_data = True
                count_col_ws = 1
                while is_data:
                    count_col_ws += 1
                    data =  self.ws.cell (row = 1, column = count_col_ws).value
                    if data == None:
                        is_data = False
                count_col_ws -=1
                
                for j in range(count_col_ws):
                    for i in range(count_row_ws):
                        if k ==0:
                            i-=1
                        self.ws_new.cell(row=ultima_linha_ws_new+i+1, column= j+1).value = self.ws.cell(row=i+2,column=j+1).value
                ultima_linha_ws_new += count_row_ws-1
                k+=1
            
            load_arquivos.excel_atualizador(self)
            

        def excel_atualizador(self):
            cliente = tela.lineEdit.text()
            banco = tela.lineEdit_2.text()
            self.cliente = cliente
            wb_twm = openpyxl.load_workbook('./arquivo_twm/arquivo_TWM_'+cliente.upper()+'.xlsx')
            wb_atualizador = openpyxl.load_workbook('./arquivo_cliente/Modelo_Campo_Customizado.xlsx')
            
            ws_atualizador_dados = wb_atualizador ['Dados Cliente']
            ws_atualizador_fatura = wb_atualizador ['Fatura']
            
            try:
                ws_twm = wb_twm['Planilha1']
            except:
                ws_twm = wb_twm['faturas']
            
            ws_atualizador_dados.cell(row = 2, column = 1).value = banco
            
            is_data = True
            count_row_ws = 1
            while is_data:
                count_row_ws += 1
                data =  ws_twm.cell (row = count_row_ws, column = 1).value
                if data == None:
                    is_data = False
            count_row_ws -=1

            i = 2
            while i <= count_row_ws:
                ws_atualizador_fatura.cell(row = i, column = 1).value = ws_twm.cell (row = i, column = 1).value
                ws_atualizador_fatura.cell(row = i, column = 2).value = 'Saneado'
                ws_atualizador_fatura.cell(row = i, column = 3).value = 'Sim'
             
                i += 1
            
            self.wb_twm = wb_twm
            self.ws_twm = ws_twm
            self.banco = banco
            
            load_arquivos.consolidado_twm(self)
            
        def consolidado_twm (self):
            self.wb_new = self.wb_new
            self.ws_new = self.ws_new
            
            self.wb_twm = self.wb_twm
            self.ws_twm = self.ws_twm
            
            is_data = True
            count_row_ws = 1
            while is_data:
                count_row_ws += 1
                data =  self.ws_new.cell (row = count_row_ws, column = 1).value
                if data == None:
                    is_data = False
            count_row_ws -=1
            
            is_data = True
            count_col_ws = 1
            while is_data:
                count_col_ws += 1
                data =  self.ws_new.cell (row = 1, column = count_col_ws).value
                if data == None:
                    is_data = False
            count_col_ws -=1
            
            is_data = True
            count_row_twm = 1
            while is_data:
                count_row_twm += 1
                data =  self.ws_twm.cell (row = count_row_twm, column = 1).value
                if data == None:
                    is_data = False
            count_row_twm -=1
            
            is_data = True
            count_col_twm = 1
            while is_data:
                count_col_twm += 1
                data =  self.ws_twm.cell (row = 1, column = count_col_twm).value
                if data == None:
                    is_data = False
            count_col_twm -=1
            
            list_col_twm = []
            for col_twm in range (count_col_twm):
                list_col_twm.append(self.ws_twm.cell(row=1, column=col_twm + 1).value)

            list_col_consolidado = []
            for col_consolidado in range (count_col_ws):
                list_col_consolidado.append(self.ws_new.cell(row=1, column=col_consolidado + 1).value)
        
            try: #Dados do TWM
                indice_conta_aglutinada = list_col_twm.index('Conta aglutinada') + 1
                indice_col_identificador = list_col_consolidado.index('Identificador') + 1
                indice_num_conta = list_col_consolidado.index('Nº da Conta') + 1 
            except: # Dados do NEXinvoice
                indice_conta_aglutinada = list_col_twm.index('CONTRATO') + 1
                indice_col_identificador = list_col_consolidado.index('FATURA') + 1
                indice_num_conta = list_col_consolidado.index('CONTRATO') + 1
            
            j = 0
            while j < len(list_col_consolidado):
                if list_col_consolidado[j] in list_col_twm:
                    indice_twm = list_col_twm.index(list_col_consolidado[j]) + 1                
                    
                    for i in range (count_row_ws-1):
                          for t in range (count_row_twm-1):          
                                if self.ws_new.cell (row = i+2, column = indice_num_conta).value == self.ws_twm.cell (row = t+2, column = indice_conta_aglutinada).value:
                                  self.ws_new.cell (row = i+2, column = j+1).value = self.ws_twm.cell (row = t+2, column = indice_twm).value
                j+=1
            #     ws_twm.delete_rows(row_delete)
            
            self.wb_twm.save('_____TWM_'+self.banco+'.xlsx')
            self.wb_new.save('_____CONSOLIDADO_COMPLETO'+self.banco+'.xlsx')
            print('Complete Consolidado...')
            
            load_arquivos.validar_consumo(self)


        def num_float(self,num):

            if num == None:
                num = float(0)

            if type(num) == str:
                string = '!.#%'
                char_rep = {k: '' for k in string}
                num = num.translate(str.maketrans(char_rep))
                num = float(num.replace(',', '.'))
            return num

        def validar_consumo(self):
            # Carregando Excel
            print()
            print('Validando consumo')
            wb_consolidado = openpyxl.load_workbook('_____CONSOLIDADO_COMPLETO' + self.banco + '.xlsx')
            wb_consolidado.save('CONSOLIDADO_Consumo' + self.banco + '.xlsx')
            ws_consolidado = wb_consolidado['Sheet']

            a = 1
            b = 0

            # Encontrar índices da coluna
            while b < a:
                try:
                    coluna = ws_consolidado.cell(row=1, column=a).value
                    if coluna == None:
                        a -= 1
                    if coluna == 'Categoria':
                        cat = a
                    if coluna == 'Subcategoria':
                        subcat = a
                    if coluna == 'Identificador':
                        idt = a
                except:
                    pass
                a += 1
                b += 1

            c = 1
            d = 0

            # Encontrar índices da linha
            while d < c:
                try:
                    linha = ws_consolidado.cell(row=c, column=idt).value
                    if linha == None:
                        c -= 1
                except:
                    pass
                c += 1
                d += 1

            i = 2

            # Filtrando Consumo
            while i <= c:
                categoria = ws_consolidado.cell(row=i, column=cat).value

                if categoria != 'Consumo':
                    ws_consolidado.delete_rows(i)
                    i -= 1
                    c -= 1
                i += 1

            l = 2

            while l <= c:
                subcategoria = ws_consolidado.cell(row=l, column=subcat).value

                if subcategoria == 'Na ponta':
                    pass
                elif subcategoria == 'Fora ponta':
                    pass
                elif subcategoria == 'TE na ponta':
                    pass
                elif subcategoria == 'TE fora ponta':
                    pass
                else:
                    ws_consolidado.delete_rows(l)
                    l -= 1
                    c -= 1
                l += 1

            wb_consolidado.save('CONSOLIDADO_Consumo' + self.banco + '.xlsx')

            # Criando tabelas dinâmicas
            tabela1 = pd.read_excel('CONSOLIDADO_Consumo' + self.banco + '.xlsx')
            tabela2 = pd.read_excel('./arquivo_twm/arquivo_TWM_' + self.cliente + '.xlsx')

            tabela_consolidado = pd.pivot_table(tabela1, index=['Identificador'], values=['Quantidade'], aggfunc=[np.sum])
            tabela_twm = pd.pivot_table(tabela2, index=['Identificador'], values=['Consumo da fatura'], aggfunc=[np.sum])

            lista_id_consolidado = tabela_consolidado.index.tolist()
            lista_consumo_consolidado = tabela_consolidado.values.tolist()
            lista_id_twm = tabela_twm.index.tolist()
            lista_consumo_twm = tabela_twm.values.tolist()

            dict_consolidado = dict(zip(lista_id_consolidado, lista_consumo_consolidado))
            dict_twm = dict(zip(lista_id_twm, lista_consumo_twm))

            count = 0

            # Encontrando consumos errados
            for key in dict_twm.keys():
                try:
                    dict_twm[key] = load_arquivos.num_float(self, dict_twm[key][0])
                    if dict_twm[key] != dict_consolidado[key][0]:
                        print(f'{key} - TWM: {dict_twm[key]} - Consolidado: {dict_consolidado[key]}')
                        count += 1
                except:
                    print(f'{key} - Não encontrado')
                    count += 1

            print(f'{count} faturas para verificar')

lista_banco_twm = ['twm_accorhotels', 'twm_segurpro', 'twm_ambev', 'twm_teste_igor', 'twm_nvivara', 'twm_vivaracsc', 'twm_mrscsc', 'twm_diego', 'twm_prosegurti', 'twm_paguemenoscsc', 'twm_fleury', 'twm_leroymerlin', 'twm_jflrealty', 'twm_kpmg', 'twm_saulo_teste', 'twm_igor_teste', 'twm_cloudkitchens', 'twm_clarke', 'twm_agibankcsc', 'twm_cpflcsc', 'twm_diagroupcsc', 'twm_deodecsc', 'twm_prosegurtaxas', 'twm_adami', 'twm_alinc', 'twm_localizaservicos', 'twm_fortlevcsc', 'twm_wilsonsons', 'twm_petrobras', 'twm_alliedshopping', 'twm_bloominbrands', 'twm_ibmsantander', 'twm_optycsc', 'twm_saintgobaincsc', 'twm_vilavitoria', 'twm_valevitoria', 'twm_gpa', 'twm_agroamazonia', 'twm_piracanjuba', 'twm_prossegurtaxas', 'twm_mrv', 'twm_demonstracao', 'twm_demo_controle', 'twm_testes_cypress', 'twm_demonstracao_homologacao', 'twm_jmalucelliti', 'twm_fca', 'twm_agiplan', 'twm_gafisa', 'twm_whirlpool', 'twm_azul', 'twm_infoglobo', 'twm_prosegurtelecom', 'twm_whirlpool_argentina', 'twm_pucminas', None, 'twm_csc', 'twm_optum', 'twm_kroton', 'comercial_twm', 'twm_cofco', 'twm_banco_modelo_validacao_backup', 'twm_rip', 'twm_comau', 'twm_servopa', 'twm_unipar', 'twm_teste_localidade', 'twm_stonefixa', 'twm_shoulder', 'twm_ale', 'twm_nimbi', 'twm_sascarm2m', None, 'twm_sulamerica', None, 'twm_zelo', 'twm_schneider', 'twm_assai', None, 'twm_brhc', 'twm_dma', 'twm_grupomarquise', 'twm_raiadrogasil', 'twm_energisa', None, 'twm_facilities_fca', 'twm_sherwin_poc', None, None, 'twm_normatel', 'twm_ligasolidaria', 'twm_voxeldigital', None, None, None, 'twm_telemont', 'twm_mrs', 'twm_amil', 'twm_jca', 'twm_preventsenior', 'twm_jmalucelli', 'twm_cnhi', 'twm_precon_engenharia', 'twm_demonstracao_cliente', None, None, 'twm_bayer', 'twm_veritas', 'twm_saintgobain', 'twm_metrorio', 'twm_zema', 'twm_fundep', 'twm_odebrecht', 'twm_carioca', 'twm_cmoc', 'twm_gross', 'twm_centralunimed', 'twm_grupo_sao_francisco', 'twm_hermes_pardini', None, 'twm_homologacao', 'twm_dhl', 'twm_embraco', 'twm_facilities_homologacao_mrv', 'twm_guiando', 'twm_boticario_escrituracao', 'twm_demo_facilities', 'twm_anglo_poc', 'twm_banco_carrefour', 'twm_integrada', 'twm_cotrijal', 'twm_hydro', 'twm_jmsegs', 'twm_duratex', 'twm_hypera', 'twm_klabin', 'twm_acsc', 'twm_leste', 'twm_alianca_homologacao', 'twm_grupo_simec', 'twm_mitsubishi', None, 'twm_ams', None, 'twm_accenture', 'twm_biolab', 'twm_rioquenteresorts', 'twm_verdecampo', None, None, 'twm_dentsuaegis', 'twm_grupo_crm', 'twm_arcosdorados', 'twm_arconic', 'twm_araujo', 'twm_facilities_homologacao_multas', 'twm_facilities_weg', 'twm_facilities_sul_america', 'twm_facilities_demonstracao', 'twm_facilities_agiplan', 'twm_facilities', 'twm_banco_modelo', None, 'twm_itatiaia', 'twm_unimedbh', 'twm_andritz_hydro', 'twm_renova', 'twm_ceabs', 'twm_chevron', 'twm_bonsucesso', 'twm_contactcenter', 'twm_continental', 'twm_csa', 'twm_dufry', 'twm_elogroup', 'twm_embrasil', 'twm_eurofarma', 'twm_expressa', 'twm_fiat', 'twm_fiesc', 'twm_fumec', 'twm_globalcob', 'twm_invepar2', 'twm_irbbrasil', 'twm_lideraviacao', 'twm_guanabara', 'twm_fenosa', None, 'twm_amc', 'twm_carrefour', 'twm_ccp', 'twm_aec', 'twm_tecnisa', 'twm_roveri', 'twm_gruposaga', 'twm_prosegur', 'twm_jsl', 'twm_dasa', 'twm_arcelor', 'twm_vitacon', 'twm_burgerkingcsc', 'twm_dpaschoal', 'twm_boticariocsc', 'twm_jll', 'twm_grpcom', 'twm_bild', 'twm_canopus', 'twm_alliedcsc', 'twm_fcaargentina', 'twm_anima', 'twm_rodobens', 'twm_direcional', 'twm_solar', 'twm_housi', 'twm_ibm', 'twm_hughes', 'twm_localiza', 'twm_localizam2m', 'twm_burgerking', 'twm_alianca', 'twm_boticario', 'twm_oillinois', 'twm_tegma', 'twm_indra', 'twm_sesc', 'twm_mipengenharia', 'twm_plenaalimentos', 'twm_terraverde', 'twm_bbm', 'twm_allied', 'twm_grupo_indigo', 'twm_magnetimarelli', 'twm_opty', 'twm_alvarezandmarsal', 'twm_fiep', 'twm_jallesmachado', 'twm_dpsp', 'twm_uniaoquimica', 'twm_hermespardini', 'twm_hirota', 'twm_linx', 'twm_stellantis', 'twm_intralot', 'twm_facilities_medgrupo', 'twm_facilities_anglo', 'twm_fortlev', 'twm_samarco', 'twm_servier', 'twm_biorad', 'twm_ebd', 'twm_bnpp', 'twm_brasilplural', 'twm_andritz', 'twm_servimed', 'twm_orguel', 'twm_alcoa', 'twm_biosev', 'twm_fmc', 'twm_gavilon', 'twm_gefco', 'twm_grupo_sada', 'twm_stone', 'twm_ache', 'twm_raizen', 'twm_michelin', 'twm_sascar', 'twm_bmg', 'twm_loreal', 'twm_sherwin', 'twm_pesa', 'twm_unimed_fortaleza', 'twm_emporiosaude', 'twm_levorin', 'twm_fiagril', 'twm_norteenergia', 'twm_gruposoma', 'twm_daki', 'twm_ribercred', 'twm_mundotelecom', 'twm_grupopetropolis', 'twm_grupoaldo', 'twm_atacadao', 'twm_aramis', 'twm_yuca', 'twm_riachuelo', 'twm_natura', 'twm_nadirfigueiredo', 'twm_unimed_sp', 'twm_vivara', 'twm_animaalunos', 'twm_mercadolivre', 'twm_parebem', 'twm_ball', 'twm_loft', 'twm_grupobmg', 'twm_grupozelo', 'twm_techint', 'twm_oncoclinicas']
lista_clientes = ['PUC', 'LOCALIZA', 'RIACHUELO', 'FLEURY', 'DAKI', 'PAGUE_MENOS', 'GRPCOM']
lista_central_twm = ['twm_pucminas', 'twm_localiza', 'twm_riachuelo', 'twm_fleury', 'twm_daki', 'twm_paguemenoscsc', 'twm_grpcom']

app = QtWidgets.QApplication([])
tela = uic.loadUi("./assets/Interface_SAM2.ui")

tela.show()

completer = QCompleter(lista_clientes)
tela.lineEdit.setCompleter(completer)

completer2 = QCompleter(lista_central_twm)
tela.lineEdit_2.setCompleter(completer2)

load_active = load_arquivos()
tela.pushButton.clicked.connect(load_active.ler_arquivo_fornecedor)
tela.pushButton_4.clicked.connect(load_active.juntar_arquivos)

sys.exit(app.exec_())





  